# -*- coding: utf-8 -*-

import json
import pickle
import os
import queue
import sys
import copy
import datetime
from threading import Thread
from time import sleep
import multiprocessing

import numpy as np
import openpyxl
import pandas as pd
from itertools import count
from teaser.project import Project
from .envelope import Envelope
from .solar import Sun
from .users import Users
from .system import BES
from .system import CES
from .plots import DemandPlots
from .optimizer import Optimizer
from .KPIs import KPIs
import districtgenerator.functions.clustering_medoid as cm


class Datahandler:
    """
    Abstract class for data handling.
    Collects data from input files, TEASER, User and Envelope.

    Attributes
    ----------
    site:
        Dict for site data, e.g. weather.
    time:
        Dict for time settings.
    district:
        List of all buildings within district.
    scenario_name:
        Name of scenario file.
    scenario:
        Scenario data.
    counter:
        Dict for counting number of equal building types.
    srcPath:
        Source path.
    filePath:
        File path.
    """

    def __init__(self, scenario_name="example", resultPath=None, scenario_file_path=None):
        """
        Constructor of Datahandler class.

        Returns
        -------
        None.
        """

        self.site = {}
        self.time = {}
        self.district = []
        self.scenario_name = scenario_name
        self.scenario = None
        self.design_building_data = {}
        self.physics = {}
        self.decentral_device_data = {}
        self.params_ehdo_technical = {}
        self.params_ehdo_model = {}
        self.central_device_data = {}
        self.ecoData = {}
        self.counter = {}
        self.srcPath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.filePath = os.path.join(self.srcPath, 'data')

        if scenario_file_path is not None:
            self.scenario_file_path = scenario_file_path
        else:
            self.scenario_file_path = os.path.join(self.filePath, 'scenarios')

        if resultPath is not None:
            self.resultPath = resultPath
        else:
            self.resultPath = os.path.join(self.srcPath, 'results')

        self.KPIs = None
        self.load_all_data()

    def load_all_data(self):
        """
        General data import from JSON files and transformation into dictionaries.

        Returns
        -------
        None.
        """

        # %% load information about of the site under consideration (used in generateEnvironment)
        # important for weather conditions
        with open(os.path.join(self.filePath, 'site_data.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                self.site[subData["name"]] = subData["value"]

        # %% load time information and requirements (used in generateEnvironment)
        # needed for data conversion into the right time format
        with open(os.path.join(self.filePath, 'time_data.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                self.time[subData["name"]] = subData["value"]

        # %% load scenario file with building information
        self.scenario = pd.read_csv(self.scenario_file_path + "/" + self.scenario_name + ".csv",
                                    header=0, delimiter=";")

        # %% load general building information
        # contains definitions and parameters that affect all buildings (used in envelope and system BES/CES)
        with open(os.path.join(self.filePath, 'design_building_data.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                self.design_building_data[subData["name"]] = subData["value"]

                # load building physics data (used in envelope and system BES/CES)
        with open(os.path.join(self.filePath, 'physics_data.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                self.physics[subData["name"]] = subData["value"]

                # Load list of possible devices (used in system BES)
        with open(os.path.join(self.filePath, 'decentral_device_data.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                self.decentral_device_data[subData["abbreviation"]] = {}
                for subsubData in subData["specifications"]:
                    self.decentral_device_data[subData["abbreviation"]][subsubData["name"]] = subsubData["value"]

        # import model parameters from json-file (used in system CES)
        with open(os.path.join(self.filePath, 'model_parameters_EHDO.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                if subData["name"] != "ref":
                    self.params_ehdo_model[subData["name"]] = subData["value"]
                else:
                    self.params_ehdo_model[subData["name"]] = {}
                    for subSubData in subData["value"]:
                        self.params_ehdo_model[subData["name"]][subSubData["name"]] = subSubData["value"]

        # load economic and ecologic data (of the district generator) (used in system CES)
        with open(os.path.join(self.filePath, 'eco_data.json')) as json_file:
            jsonData = json.load(json_file)
            for subData in jsonData:
                self.ecoData[subData["name"]] = subData["value"]

        with open(os.path.join(self.filePath, 'central_device_data.json')) as json_file:
            self.central_device_data = json.load(json_file)

    def select_plz_data(self):
        """
        Select the closest TRY weather station for the location of the postal code.

        Returns
        -------
        weatherdatafile_location: int
            Location of the TRY weather station in lambert projection.
        """

        # Try to find the location of the postal code and matched TRY weather station
        try:
            workbook = openpyxl.load_workbook(self.filePath + "/plz_geocoord_matched.xlsx")
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if self.site["zip"] == str(row[0]):
                    weatherdatafile = row[3]
                    self.site["Location"] = weatherdatafile[8:-9]
                    break
            else:
                # If postal code cannot be found: Message and select weather data file from Aachen
                raise ValueError("Postal code cannot be found")


        except Exception as e:
            # If postal code cannot be found: Message and select weathter data file from Aachen
            print("Postal code cannot be found, location changed to Aachen")
            self.site["zip"] = "52064"
            self.site["Location"] = 507755060854
            """  
                Add new weatherdatafile_location, if you want an individual location: 
                Files can be found here: https://www.dwd.de/DE/leistungen/testreferenzjahre/testreferenzjahre.html 
                Every file has to be stored in the folder reffering to the correct Year and season in the subfolders of '\districtgenerator\data\weather\ 
                Example: TRY2015_507755060854_Wint.dat has to be stored in '\districtgenerator\data\weather\TRY_2015_Winter' 
                Uncomment the following line  
            """
            # weatherdatafile_location = 507755060854

    def generateEnvironment(self):
        """
        Load physical district environment - site and weather.

        Returns
        -------
        None.
        """
        # %% load first day of the year
        if self.site["TRYYear"] == "TRY2015":
            first_row = 35
        elif self.site["TRYYear"] == "TRY2045":
            first_row = 37

        self.select_plz_data()
        # load weather data
        # select the correct file depending on the TRY weather station location
        weatherData = np.loadtxt(
            os.path.join(self.filePath, "weather", "TRY_" + self.site["TRYYear"][-4:] + "_" + self.site["TRYType"])
            + "/"
            + self.site["TRYYear"] + "_"
            + str(self.site["Location"]) + "_" + str(self.site["TRYType"])
            + ".dat",
            skiprows=first_row - 1)

        """
        # Use this function to load old TRY-weather data
        weatherData = np.loadtxt(os.path.join(self.filePath, 'weather')
                                 + "/"
                                 + self.site["TRYYear"] + "_Zone"
                                 + str(self.site["climateZone"]) + "_"
                                 + self.site["TRYType"] + ".txt",
                                 skiprows=first_row - 1)"""

        # weather data starts with 1st january at 1:00 am.
        # Add data point for 0:00 am to be able to perform interpolation.
        weatherData_temp = weatherData[-1:, :]
        weatherData = np.append(weatherData_temp, weatherData, axis=0)

        # get weather data of interest
        [temp_sunDirect, temp_sunDiff, temp_temp, temp_wind] = \
            [weatherData[:, 12], weatherData[:, 13], weatherData[:, 5], weatherData[:, 8]]

        self.time["timeSteps"] = int(self.time["dataLength"] / self.time["timeResolution"])

        # load the holidays
        if self.site["TRYYear"] == "TRY2015":
            self.time["holidays"] = self.time["holidays2015"]
        elif self.site["TRYYear"] == "TRY2045":
            self.time["holidays"] = self.time["holidays2045"]

        # interpolate input data to achieve required data resolution
        # transformation from values for points in time to values for time intervals
        self.site["SunDirect"] = np.interp(np.arange(0, self.time["dataLength"] + 1, self.time["timeResolution"]),
                                           np.arange(0, self.time["dataLength"] + 1, self.time["dataResolution"]),
                                           temp_sunDirect)[0:-1]
        self.site["SunDiffuse"] = np.interp(np.arange(0, self.time["dataLength"] + 1, self.time["timeResolution"]),
                                            np.arange(0, self.time["dataLength"] + 1, self.time["dataResolution"]),
                                            temp_sunDiff)[0:-1]
        self.site["T_e"] = np.interp(np.arange(0, self.time["dataLength"] + 1, self.time["timeResolution"]),
                                     np.arange(0, self.time["dataLength"] + 1, self.time["dataResolution"]),
                                     temp_temp)[0:-1]
        self.site["wind_speed"] = np.interp(np.arange(0, self.time["dataLength"] + 1, self.time["timeResolution"]),
                                            np.arange(0, self.time["dataLength"] + 1, self.time["dataResolution"]),
                                            temp_wind)[0:-1]

        self.site["SunTotal"] = self.site["SunDirect"] + self.site["SunDiffuse"]

        # Load other site-dependent values based on DIN/TS 12831-1:2020-04
        srcPath = os.path.dirname(os.path.abspath(__file__))
        filePath = os.path.join(os.path.dirname(srcPath), 'data', 'site_data.txt')
        site_data = pd.read_csv(filePath, delimiter='\t', dtype={'Zip': str})

        # Filter data for the specific zip code
        filtered_data = site_data[site_data['Zip'] == self.site["zip"]]

        # extract the needed values
        self.site["altitude"] = filtered_data.iloc[0]['Altitude']
        self.site["location"] = [filtered_data.iloc[0]['Latitude'], filtered_data.iloc[0]['Longitude']]
        self.site["T_ne"] = filtered_data.iloc[0][
            'T_ne']  # norm outside temperature for calculating the design heat load
        self.site["T_me"] = filtered_data.iloc[0][
            'T_me']  # mean annual temperature for calculating the design heat load

        # Calculate solar irradiance per surface direction - S, W, N, E, Roof represented by angles gamma and beta
        global sun
        sun = Sun(filePath=self.filePath)
        self.site["SunRad"] = sun.getSolarGains(initialTime=0,
                                                timeDiscretization=self.time["timeResolution"],
                                                timeSteps=self.time["timeSteps"],
                                                timeZone=self.site["timeZone"],
                                                location=self.site["location"],
                                                altitude=self.site["altitude"],
                                                beta=[90, 90, 90, 90, 0],
                                                gamma=[0, 90, 180, 270, 0],
                                                beamRadiation=self.site["SunDirect"],
                                                diffuseRadiation=self.site["SunDiffuse"],
                                                albedo=self.site["albedo"])

    def initializeBuildings(self):
        """
        Fill district with buildings from scenario file.

        Parameters
        ----------
        scenario_name: string, optional
            Name of scenario file to be read. The default is 'example'.

        Returns
        -------
        None.
        """
        duration = datetime.timedelta(minutes=1)
        num_sfh = 0
        num_mfh = 0
        name_pool = []

        # initialize buildings for scenario
        # loop over all buildings
        for id in self.scenario["id"]:
            try:
                # Check if ID is a number
                try:
                    # Try to convert id to float to check if it's numeric
                    float(id)
                except (ValueError, TypeError):
                    raise ValueError(f"Building ID '{id}' is not a number")

                # Create empty dict for observed building
                building = {}

                # Store features of the observed building
                building["buildingFeatures"] = self.scenario.loc[id]

                # %% Create unique building name
                # needed for loading and storing data with unique name
                # name is composed of building id, and building type
                name = str(id) + "_" + building["buildingFeatures"]["building"]

                # Check if the name is already in the district
                if name in name_pool:
                    raise ValueError(f"Building name '{name}' is not unique. ID collision detected.")

                name_pool.append(name)

                # Assign the unique name to the building
                building["unique_name"] = name

                # Append building to district
                self.district.append(building)

                # Count number of builidings to predict the approximate calculation time
                if building["buildingFeatures"]["building"] == 'SFH' or building["buildingFeatures"][
                    "building"] == 'TH':
                    num_sfh += 1
                elif building["buildingFeatures"]["building"] == 'MFH' or building["buildingFeatures"][
                    "building"] == 'AB':
                    num_mfh += 1
            except ValueError as e:
                # Handle the case where we have a duplicate name
                print(f"Error: {e}")
                print(f"The building ID must be a unique number to ensure proper identification and data tracking.")
                print(f"Building with ID {id} will be skipped and not added to the district")
                continue
            except Exception as e:
                # Handle any other unexpected errors
                print(f"Unexpected error processing building ID {id}: {e}")
                print(f"Building with ID {id} will be skipped and not added to the district.")
                continue

        # Calculate calculation time for the whole district generation
        duration += datetime.timedelta(seconds=3 * num_sfh + 12 * num_mfh)
        print("This calculation will take about " + str(duration) + " .")

    def generateBuildings(self):
        """
        Load building envelope and user data.

        Returns
        -------
        None.
        """

        # %% load general building information
        # contains definitions and parameters that affect all buildings
        bldgs = self.design_building_data

        # %% create TEASER project
        # create one project for the whole district
        prj = Project(load_data=True)
        prj.name = self.scenario_name

        for building in self.district:

            # convert short names into designation needed for TEASER
            building_type = \
                bldgs["buildings_long"][bldgs["buildings_short"].index(building["buildingFeatures"]["building"])]
            retrofit_level = \
                bldgs["retrofit_long"][bldgs["retrofit_short"].index(building["buildingFeatures"]["retrofit"])]

            # add buildings to TEASER project
            prj.add_residential(method='tabula_de',
                                usage=building_type,
                                name="ResidentialBuildingTabula",
                                year_of_construction=building["buildingFeatures"]["year"],
                                number_of_floors=3,
                                height_of_floors=3.125,
                                net_leased_area=building["buildingFeatures"]["area"],
                                construction_type=retrofit_level)

            # %% create envelope object
            # containing all physical data of the envelope
            building["envelope"] = Envelope(prj=prj,
                                            building_params=building["buildingFeatures"],
                                            construction_type=retrofit_level,
                                            physics=self.physics,
                                            design_building_data=self.design_building_data,
                                            file_path=self.filePath)

            # %% create user object
            # containing number occupants, electricity demand,...
            building["user"] = Users(building=building["buildingFeatures"]["building"],
                                     area=building["buildingFeatures"]["area"])

            # %% calculate design heat loads
            # at norm outside temperature
            building["envelope"].heatload = building["envelope"].calcHeatLoad(site=self.site, method="design")
            # at bivalent temperature
            building["envelope"].bivalent = building["envelope"].calcHeatLoad(site=self.site, method="bivalent")
            # at heating limit temperature
            building["envelope"].heatlimit = building["envelope"].calcHeatLoad(site=self.site, method="heatlimit")
            # for drinking hot water
            if building["user"].building in {"SFH", "MFH", "TH", "AB"}:
                building["dhwload"] = bldgs["dhwload"][bldgs["buildings_short"].index(building["user"].building)] * \
                                      building["user"].nb_flats
            else:
                building["dhwload"] = bldgs["dhwload"][bldgs["buildings_short"].index(building["user"].building)] * \
                                      building["user"].nb_main_rooms

            index = bldgs["buildings_short"].index(building["buildingFeatures"]["building"])
            building["buildingFeatures"]["mean_drawoff_dhw"] = bldgs["mean_drawoff_vol_per_day"][index]

    def generateDemands(self, calcUserProfiles=True, saveUserProfiles=True):
        """
        Multiprocessing-Version von generateDemands
        """
        cpu_count = multiprocessing.cpu_count()
        args_list = [(self, building, calcUserProfiles, saveUserProfiles) for building in self.district]
        # cpu_count = 4

        with multiprocessing.Pool(processes=cpu_count) as pool:
            results = pool.map(generate_demands_worker_wrapper, args_list)

        for result in results:
            building = next(b for b in self.district if b["unique_name"] == result["unique_name"])
            building["user"].elec = result["elec"]
            building["user"].dhw = result["dhw"]
            building["user"].cooling = result["cooling"]
            building["user"].heat = result["heating"]
            building["user"].occ = result["occ"],
            building["user"].car = result["car"]
            building["user"].gains = result["gains"]
            building["user"].nb_flats = result["nb_flats"]
            building["user"].nb_occ = result["nb_occ"]
            building["envelope"] = result["envelope"]
            building["clusteringData"] = result["clusteringData"]
            building_features = building["buildingFeatures"].copy()
            building_features["night_setback"] = result["night_setback"]
            building["buildingFeatures"] = building_features

        print("Finished generating demands with multiprocessing!")


    def generate_demands_worker(self, building, calcUserProfiles, saveUserProfiles):
        """
        :param building:
        :param calcUserProfiles: bool
            True: calculate new user profiles.
            False: load user profiles from file.
            The default is True.
        :param saveUserProfiles: bool
            True for saving calculated user profiles in workspace (Only taken into account if calcUserProfile is True).
            The default is True.
        """
        print(f'starting {building["unique_name"]}')

        # calculate or load user profiles
        if calcUserProfiles:
            building["user"].calcProfiles(site=self.site,
                                          holidays=self.time["holidays"],
                                          time_resolution=self.time["timeResolution"],
                                          time_horizon=self.time["dataLength"],
                                          building=building,
                                          path=os.path.join(self.resultPath, 'demands'))

            if saveUserProfiles:
                self.saveProfiles(name=building["unique_name"],
                                  elec=building["user"].elec,
                                  dhw=building["user"].dhw,
                                  occ=building["user"].occ,
                                  gains=building["user"].gains,
                                  car=building["user"].car,
                                  nb_flats=building["user"].nb_flats,
                                  nb_occ=building["user"].nb_occ,
                                  heatload=building["envelope"].heatload,
                                  bivalent=building["envelope"].bivalent,
                                  heatlimit=building["envelope"].heatlimit,
                                  path=os.path.join(self.resultPath, 'demands'))
                # building["user"].saveProfiles(building["unique_name"], building["envelope"], os.path.join(self.resultPath, 'demands'))

            print("Calculate demands of building " + building["unique_name"])

        else:
            (building["user"].elec, building["user"].dhw,
             building["user"].occ, building["user"].gains,
             building["user"].car, building["user"].nb_flats,
             building["user"].nb_occ, building["envelope"].heatload,
             building["envelope"].bivalent,
             building["envelope"].heatlimit) = self.loadProfiles(building["unique_name"],
                                                                 os.path.join(self.resultPath, 'demands'))
            # building["user"].loadProfiles(building["unique_name"], os.path.join(self.resultPath, 'demands'))
            print("Load demands of building " + building["unique_name"])

        # check if EV exist
        building["clusteringData"] = {
            "potentialEV": copy.deepcopy(building["user"].car)
        }
        building["user"].car *= building["buildingFeatures"]["EV"]

        building["envelope"].calcNormativeProperties(self.site["SunRad"], building["user"].gains)

        night_setback = building["buildingFeatures"]["night_setback"]

        # calculate or load heating profiles
        if calcUserProfiles:
            building["user"].calcHeatingProfile(site=self.site,
                                                envelope=building["envelope"],
                                                night_setback=night_setback,
                                                holidays=self.time["holidays"],
                                                time_resolution=self.time["timeResolution"]
                                                )

            if saveUserProfiles:
                self.saveHeatingProfile(heat=building["user"].heat,
                                        cooling=building["user"].cooling,
                                        name=building["unique_name"],
                                        path=os.path.join(self.resultPath, 'demands'))
                # building["user"].saveHeatingProfile(building["unique_name"], os.path.join(self.resultPath, 'demands'))
        else:
            heat, cooling = self.loadHeatingProfiles(name=building["unique_name"],
                                                     path=os.path.join(self.resultPath, 'demands'))
            building["user"].heat = heat
            building["user"].cooling = cooling
        print(f'done {building["unique_name"]}')

    def generateDistrictComplete(self, calcUserProfiles=True, saveUserProfiles=True,
                                 saveGenProfiles=True, designDevs=False, clustering=False, optimization=False):
        """
        All in one solution for district and demand generation.

        Parameters
        ----------

        calcUserProfiles: bool, optional
            True: calculate new user profiles.
            False: load user profiles from file.
            The default is True.
        saveUserProfiles: bool, optional
            True for saving calculated user profiles in workspace (Only taken into account if calcUserProfile is True).
            The default is True.
        fileName_centralSystems : string, optional
            File name of the CSV-file that will be loaded. The default is "central_devices_test".
        saveGenProfiles: bool, optional
            Decision if generation profiles of designed devices will be saved. Just relevant if 'designDevs=True'.
            The default is True.
        designDevs: bool, optional
            Decision if devices will be designed. The default is False.
        clustering: bool, optional
            Decision if profiles will be clustered. The default is False.
        optimization: bool, optional
            Decision if the operation costs for each cluster will be optimized. The default is False.

        Returns
        -------
        None.
        """

        self.initializeBuildings()
        self.generateEnvironment()
        self.generateBuildings()
        self.generateDemands(calcUserProfiles, saveUserProfiles)
        if designDevs:
            self.designDevicesComplete(saveGenProfiles)
        if clustering:
            if designDevs:
                self.clusterProfiles()
            else:
                print("Clustering is not possible without the design of energy conversion devices!")
        if optimization:
            if designDevs and clustering:
                self.optimizationClusters()
            else:
                print("Optimization is not possible without clustering and the design of energy conversion devices!")

    def saveProfiles(self, name, elec, dhw, occ, gains, car, nb_flats, nb_occ, heatload, bivalent, heatlimit, path):
        """
        Save profiles to csv.

        Parameters
        ----------
        unique_name : string
            Unique building name.
        path : string
            Results path.

        Returns
        -------
        None.
        """

        other_data = [nb_flats, str(nb_occ)[1:-1], heatload, bivalent, heatlimit]

        data_dict = {
            'elec': (elec, "Electricity demand in W"),
            'dhw': (dhw, "Drinking hot water in W"),
            'occ': (occ, "Occupancy of persons"),
            'gains': (gains, "Internal gains in W"),
            'car': (car, "Electricity demand of EV in W"),
            'other': (other_data, "Number of flats , "
                                  "Number of occupants, "
                                  "Design heat load in W, "
                                  "Bivalent heat load in W, "
                                  "Heat limit heat load in W"),
        }

        excel_file = os.path.join(path, name + '.xlsx')
        with pd.ExcelWriter(excel_file, engine='xlsxwriter') as writer:
            for sheet_name, (data, header) in data_dict.items():
                df = pd.DataFrame(data)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=header)

    def saveHeatingProfile(self, heat, cooling, name, path):
        """
        Save heating demand to csv.

        Parameters
        ----------
        unique_name : string
            Unique building name.
        path : string
            Results path.

        Returns
        -------
        None.
        """

        excel_file = os.path.join(path, name + '.xlsx')
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            cooling_df = pd.DataFrame(cooling)
            heating_df = pd.DataFrame(heat)
            cooling_df.to_excel(writer, sheet_name='cooling', index=False, header='Cooling in W')
            heating_df.to_excel(writer, sheet_name='heating', index=False, header='Heating in W')

    def loadProfiles(self, name, path):
        """
        Load profiles from csv.

        Parameters
        ----------
        unique_name : string
            Unique building name.
        path : string
            Results path.

        Returns
        -------
        None.
        """

        excel_file = os.path.join(path, name + '.xlsx')
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        def load_sheet_to_numpy(workbook, sheet_name):
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row[0])
            return np.array(data)

        elec = load_sheet_to_numpy(workbook, 'elec')
        dhw = load_sheet_to_numpy(workbook, 'dhw')
        occ = load_sheet_to_numpy(workbook, 'occ')
        gains = load_sheet_to_numpy(workbook, 'gains')
        car = load_sheet_to_numpy(workbook, 'car')
        other_data = load_sheet_to_numpy(workbook, 'other')
        nb_flats = int(other_data[0])
        nb_occ = np.fromstring(other_data[1], dtype=int, sep=',')
        heatload = float(other_data[2])
        bivalent = float(other_data[3])
        heatlimit = float(other_data[4])

        workbook.close()

        return elec, dhw, occ, gains, car, nb_flats, nb_occ, heatload, bivalent, heatlimit

    def loadHeatingProfiles(self, name, path):
        """
        Load profiles from csv.

        Parameters
        ----------
        unique_name : string
            Unique building name.
        path : string
            Results path.

        Returns
        -------
        None.
        """

        excel_file = os.path.join(path, name + '.xlsx')
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        def load_sheet_to_numpy(workbook, sheet_name):
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row[0])
            return np.array(data)

        heat = load_sheet_to_numpy(workbook, 'heating')
        cooling = load_sheet_to_numpy(workbook, 'cooling')

        workbook.close()

        return heat, cooling

    def designDecentralDevices(self, saveGenerationProfiles=True):
        """
        Calculate capacities, generation profiles of renewable energies and EV load profiles for decentral devices.

        Parameters
        ----------
        saveGenerationProfiles : bool, optional
            True: save decentral PV and STC profiles as CSV-file.
            False: don't save decentral PV and STC profiles as CSV-file.
            The default is True.

        Returns
        -------
        None.
        """

        for building in self.district:

            # %% create building energy system object
            # get capacities of all possible devices
            bes_obj = BES(physics=self.physics,
                          decentral_device_data=self.decentral_device_data,
                          design_building_data=self.design_building_data,
                          file_path=self.filePath)
            building["capacities"] = bes_obj.designECS(building, self.site)

            # calculate theoretical PV and STC generation
            potentialPV, potentialSTC = \
                sun.calcPVAndSTCProfile(time=self.time,
                                        site=self.site,
                                        area_roof=building["envelope"].A["opaque"]["roof"],
                                        # In Germany, this is a roof pitch between 30 and 35 degrees
                                        beta=[35],
                                        # surface azimuth angles (Orientation to the south: 0°)
                                        gamma=[building["buildingFeatures"]["gamma_PV"]],
                                        usageFactorPV=building["buildingFeatures"]["f_PV"],
                                        usageFactorSTC=building["buildingFeatures"]["f_STC"])

            # assign real PV generation to building
            building["generationPV"] = potentialPV * building["buildingFeatures"]["PV"]

            # assign real STC generation to building
            building["generationSTC"] = potentialSTC * building["buildingFeatures"]["STC"]

            # clustering data
            building["clusteringData"]["potentialPV"] = potentialPV
            building["clusteringData"]["potentialSTC"] = potentialSTC

            # optionally save generation profiles
            if saveGenerationProfiles == True:
                np.savetxt(os.path.join(self.resultPath, 'generation')
                           + '/decentralPV_' + building["unique_name"] + '.csv',
                           building["generationPV"],
                           delimiter=',')
                np.savetxt(os.path.join(self.resultPath, 'generation')
                           + '/decentralSTC_' + building["unique_name"] + '.csv',
                           building["generationSTC"],
                           delimiter=',')

    def designCentralDevices(self, saveGenerationProfiles):
        """
        Calculate capacities and generation profiles of renewable energies for central devices.

        Parameters
        ----------
        saveGenerationProfiles : bool, optional
            True: save central PV, STC and WT profiles as CSV-file.
            False: don't save central PV, STC and WT profiles as CSV-file.
            The default is True.


        Returns
        -------
        None.
        """
        # initialization
        self.centralDevices = {}

        # initialize central energy system object
        self.centralDevices["ces_obj"] = CES()

        # dimensioning of central devices
        self.centralDevices["capacities"] = self.centralDevices["ces_obj"].designCES(self)

        # calculate theoretical PV, STC and Wind generation
        self.centralDevices["generation"] = {}
        (self.centralDevices["generation"]["PV"], self.centralDevices["generation"]["STC"],
         self.centralDevices["generation"]["Wind"]) = self.centralDevices["ces_obj"].generation(self)

        # optionally save generation profiles
        if saveGenerationProfiles == True:
            np.savetxt(os.path.join(self.resultPath, 'generation', 'centralPV.csv'),
                       self.centralDevices["generation"]["PV"],
                       delimiter=',')
            np.savetxt(os.path.join(self.resultPath, 'generation', 'centralSTC.csv'),
                       self.centralDevices["generation"]["STC"],
                       delimiter=',')
            np.savetxt(os.path.join(self.resultPath, 'generation', 'centralWind.csv'),
                       self.centralDevices["generation"]["Wind"],
                       delimiter=',')

    def designDevicesComplete(self, saveGenerationProfiles=True):
        """
        Design decentral and central devices.

        Parameters
        ----------
        fileName_centralSystems : string, optional
            File name of the CSV-file that will be loaded. The default is "central_devices_test".
        saveGenerationProfiles : bool, optional
            Decision if generation profiles of designed devices will be saved. The default is True.

        Returns
        -------
        None.
        """

        self.designDecentralDevices(saveGenerationProfiles)
        self.designCentralDevices(saveGenerationProfiles)

    def clusterProfiles(self, centralEnergySupply):
        """
        Perform time series aggregation for profiles by using the k-medoids clustering algorithm.

        Returns
        -------
        None.
        """

        # calculate cluster time horizon
        initialArrayLenght = (self.time["clusterLength"] / self.time["timeResolution"])
        lenghtArray = initialArrayLenght
        while lenghtArray <= len(self.site["T_e"]):
            lenghtArray += initialArrayLenght
        lenghtArray -= initialArrayLenght
        lenghtArray = int(lenghtArray)

        # adjust profiles with calculated array length
        adjProfiles = {}
        # loop over buildings
        for id in self.scenario["id"]:
            adjProfiles[id] = {}
            adjProfiles[id]["elec"] = self.district[id]["user"].elec[0:lenghtArray]
            adjProfiles[id]["dhw"] = self.district[id]["user"].dhw[0:lenghtArray]
            adjProfiles[id]["heat"] = self.district[id]["user"].heat[0:lenghtArray]
            adjProfiles[id]["cooling"] = self.district[id]["user"].cooling[0:lenghtArray]
        if centralEnergySupply == False:
            for id in self.scenario["id"]:
                adjProfiles[id]["car"] = self.district[id]["user"].car[0:lenghtArray]
                adjProfiles[id]["generationPV"] = self.district[id]["generationPV"][0:lenghtArray]
                adjProfiles[id]["generationSTC"] = self.district[id]["generationSTC"][0:lenghtArray]

        if centralEnergySupply == True:
            if self.centralDevices["capacities"]["power_kW"]["WT"] > 0:
                existence_centralWT = 1
                adjProfiles["generationCentralWT"] = self.centralDevices["generation"]["Wind"][0:lenghtArray]
            else:
                # no central WT exists; but array with just zeros leads to problem while clustering
                existence_centralWT = 0
                adjProfiles["generationCentralWT"] = np.ones(lenghtArray) * sys.float_info.epsilon
            if self.centralDevices["capacities"]["power_kW"]["PV"] > 0:
                existence_centralPV = 1
                adjProfiles["generationCentralPV"] = self.centralDevices["generation"]["PV"][0:lenghtArray]
            else:
                # no central PV exists; but array with just zeros leads to problem while clustering
                existence_centralPV = 0
                adjProfiles["generationCentralPV"] = np.ones(lenghtArray) * sys.float_info.epsilon
            if self.centralDevices["capacities"]["heat_kW"]["STC"] > 0:
                existence_centralSTC = 1
                adjProfiles["generationCentralSTC"] = self.centralDevices["generation"]["STC"][0:lenghtArray]
            else:
                # no central STC exists; but array with just zeros leads to problem while clustering
                existence_centralSTC = 0
                adjProfiles["generationCentralSTC"] = np.ones(lenghtArray) * sys.float_info.epsilon

        # wind speed and ambient temperature
        adjProfiles["T_e"] = self.site["T_e"][0:lenghtArray]

        # Prepare clustering
        inputsClustering = []
        # loop over buildings
        for id in self.scenario["id"]:
            inputsClustering.append(adjProfiles[id]["elec"])
            inputsClustering.append(adjProfiles[id]["dhw"])
            inputsClustering.append(adjProfiles[id]["heat"])
            inputsClustering.append(adjProfiles[id]["cooling"])
            if centralEnergySupply == False:
                inputsClustering.append(adjProfiles[id]["car"])
                inputsClustering.append(adjProfiles[id]["generationPV"])
                inputsClustering.append(adjProfiles[id]["generationSTC"])
        # ambient temperature
        inputsClustering.append(adjProfiles["T_e"])
        if centralEnergySupply == True:
            # central renewable generation
            inputsClustering.append(adjProfiles["generationCentralWT"])
            inputsClustering.append(adjProfiles["generationCentralPV"])
            inputsClustering.append(adjProfiles["generationCentralSTC"])

        # weights for clustering algorithm indicating the focus onto this profile
        weights = np.ones(len(inputsClustering))
        # higher weight for outdoor temperature (should at least have the same weight as number of buildings)
        weights[-1] = len(self.scenario["id"])

        # Perform clustering
        (newProfiles, nc, y, z, transfProfiles) = cm.cluster(np.array(inputsClustering),
                                                             number_clusters=self.time["clusterNumber"],
                                                             len_cluster=int(initialArrayLenght),
                                                             weights=weights)

        # safe clustered profiles of all buildings
        for id in self.scenario["id"]:
            if centralEnergySupply == False:
                index_house = int(7)  # number of profiles per building
            else:
                index_house = int(4)
            self.district[id]["user"].elec_cluster = newProfiles[index_house * id]
            self.district[id]["user"].dhw_cluster = newProfiles[index_house * id + 1]
            self.district[id]["user"].heat_cluster = newProfiles[index_house * id + 2]
            self.district[id]["user"].cooling_cluster = newProfiles[index_house * id + 3]
            if centralEnergySupply == False:
                self.district[id]["user"].car_cluster = newProfiles[index_house * id + 4] * self.scenario.loc[id]["EV"]
                self.district[id]["generationPV_cluster"] = newProfiles[index_house * id + 5] \
                                                            * self.district[id]["buildingFeatures"]["PV"]
                self.district[id]["generationSTC_cluster"] = newProfiles[index_house * id + 6] \
                                                             * self.district[id]["buildingFeatures"]["STC"]

        if centralEnergySupply == True:
            self.site["T_e_cluster"] = newProfiles[-4]
            self.centralDevices["generation"]["Wind_cluster"] = newProfiles[-3]
            self.centralDevices["generation"]["PV_cluster"] = newProfiles[-2]
            self.centralDevices["generation"]["STC_cluster"] = newProfiles[-1]
        else:
            self.site["T_e_cluster"] = newProfiles[-1]

        # clusters
        self.clusters = []
        for i in range(len(y)):
            if y[i] != 0:
                self.clusters.append(i)

        # clusters and their assigned nodes (days/weeks/etc)
        self.clusterAssignments = {}
        for c in self.clusters:
            self.clusterAssignments[c] = []
            temp = z[c]
            for i in range(len(temp)):
                if temp[i] == 1:
                    self.clusterAssignments[c].append(i)

        # weights indicating how often a cluster appears
        self.clusterWeights = {}
        for c in self.clusters:
            self.clusterWeights[c] = len(self.clusterAssignments[c])

    def saveDistrict(self):
        """
        Save district dict as pickle file.

        Returns
        -------
        None.
        """

        with open(self.resultPath + "/" + self.scenario_name + ".p", 'wb') as fp:
            pickle.dump(self.district, fp, protocol=pickle.HIGHEST_PROTOCOL)

    def loadDistrict(self, scenario_name='example'):
        """
        Load district dict from pickle file.

        Parameters
        ----------
        scenario_name : string, optional
            Name of district file to be read. The default is 'example'.

        Returns
        -------
        None.
        """

        self.scenario_name = scenario_name

        with open(self.resultPath + "/" + self.scenario_name + ".p", 'rb') as fp:
            self.district = pickle.load(fp)

    def plot(self, mode='default', initialTime=0, timeHorizon=31536000, savePlots=True, timeStamp=False, show=False):
        """
        Create plots of the energy consumption and generation.

        Parameters
        ----------
        mode : string, optional
            Choose a single plot or show all of them as default. The default is 'default'.
            Possible modes are:
            ['elec', 'dhw', 'gains', 'occ', 'car', 'heating', 'pv', 'stc', 'electricityDemand', 'heatDemand'].
        initialTime : integer, optional
            Start of the plot in seconds from the beginning of the year. The default is 0.
        timeHorizon : integer, optional
            Length of the time horizon that is plotted in seconds. The default is 31536000 (what equals one year).
        savePlots : boolean, optional
            Decision if plots are saved under results/plots/. The default is True.
        timeStamp : boolean, optional
            Decision if saved plots get a unique name by adding a time stamp. The default is False.
        show : boolean, optional
            Decision if saved plots are presented directly to the user. The default is False.

        Returns
        -------
        None.
        """

        # initialize plots and prepare data for plotting
        demandPlots = DemandPlots(resultPath=self.resultPath)
        demandPlots.preparePlots(self)

        # check which resolution for plots is used
        if initialTime == 0 and timeHorizon == 31536000:
            plotResolution = 'monthly'
        else:
            plotResolution = 'stepwise'

        # the selection of possible plots
        plotTypes = \
            ['elec', 'dhw', 'gains', 'occ', 'car', 'heating', 'pv', 'stc', 'electricityDemand', 'heatDemand', 'wt']

        if mode == 'default':
            # create all default plots
            demandPlots.defaultPlots(plotResolution, initialTime=initialTime, timeHorizon=timeHorizon,
                                     savePlots=savePlots, timeStamp=timeStamp, show=show)
        elif mode in plotTypes:
            # create a plot
            demandPlots.onePlot(plotType=mode, plotResolution=plotResolution, initialTime=initialTime,
                                timeHorizon=timeHorizon, savePlots=savePlots, timeStamp=timeStamp, show=show)
        else:
            # print massage that input is not valid
            print('\n Selected plot mode is not valid. So no plot could de generated. \n')

    def optimizationClusters(self):
        """
        Optimize the operation costs for each cluster.

        Returns
        -------
        None.
        """
        optiData = {}

        # initialize result list for all clusters
        self.resultsOptimization = []

        for cluster in range(self.time["clusterNumber"]):
            # optimize operating costs of the district for current cluster
            self.optimizer = Optimizer(self, cluster)
            results_temp = self.optimizer.run_cen_opti()

            # save results as attribute
            self.resultsOptimization.append(results_temp)

    def calulateKPIs(self):
        """
        Calculate key performance indicators (KPIs).

        Returns
        -------
        None.
        """

        # initialize KPI class
        self.KPIs = KPIs(self)
        # calculate KPIs
        self.KPIs.calculateAllKPIs(self)


def generate_demands_worker_wrapper(args):
    """
    Wrapper-Funktion außerhalb der Klasse, da multiprocessing pickling benötigt.
    Args enthält (building, calcUserProfiles, saveUserProfiles, andere Parameter)
    """
    self_ref, building, calcUserProfiles, saveUserProfiles = args
    self_ref.generate_demands_worker(building, calcUserProfiles, saveUserProfiles)

    result = {
        "unique_name": building["unique_name"],
        "elec": building["user"].elec,
        'dhw': building["user"].dhw,
        'cooling': building["user"].cooling,
        'heating': building["user"].heat,
        'occ': building["user"].occ,
        'car': building["user"].car,
        'gains': building["user"].gains,
        'nb_flats': building["user"].nb_flats,
        'nb_occ': building["user"].nb_occ,
        'envelope': building["envelope"],
        'clusteringData': building["clusteringData"],
        'night_setback': building["buildingFeatures"]["night_setback"].iloc[0] if hasattr(building["buildingFeatures"]["night_setback"], 'iloc') else building["buildingFeatures"]["night_setback"],
    }

    return result
